import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import altair as alt
import numpy as np
#from streamlit_extras.colored_header import colored_header
import matplotlib.pyplot as plt
from streamlit_echarts import st_echarts 
# Fecha actual
FECHA = datetime.now().strftime('%d-%m-%y')

# Configuración de la página
st.set_page_config(
    page_title="Control de Proyectos",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded")

# Ruta al archivo de Excel
BD = './BD.xlsx'

# Cargar el libro de trabajo y las hojas disponibles
TEMPORADAS = load_workbook(BD, read_only=True).sheetnames

# Configuración de la barra lateral
st.sidebar.header('Proyectos :anchor:')
selector_temporada = st.sidebar.selectbox("Seleccione la temporada:", TEMPORADAS, index=0)
df_proyecto = pd.read_excel(BD, sheet_name=selector_temporada)

# Mostrar opciones para seleccionar proyectos
selector_proyecto = st.sidebar.multiselect("Seleccione proyectos:", df_proyecto['Proyecto'].drop_duplicates())
selector_materiales = st.sidebar.select_slider(
    "Materiales:",
    options=["Estimado","Despachado",])

if selector_materiales=="Estimado":
    MAT = 'MAT Estimado'
    PESO = 'Peso estimado(kg)'
    CTD = 'Cantidad'
else:
    MAT = 'MAT Despachado'
    PESO = 'Peso(kg)'
    CTD = 'Cantidad tomada'

precio_dolar = st.sidebar.slider("Precio del dólar:",min_value=3.00, max_value=4.00, value=3.75, step=0.05,)

# Título de la aplicación
st.title(":ship: Control de proyectos")

# Ruta al archivo de Excel
UTI = './'+selector_temporada+'/UTI.xlsx'
df_UTI = pd.read_excel(UTI, sheet_name="Sheet1")

# Reemplazar valores N/A (NaN) con ceros
df_UTI = df_UTI.fillna(0)

REDI = './'+selector_temporada+'/REDI.xlsx'
df_REDI = pd.read_excel(REDI, sheet_name="Sheet1")

#SUMA DE ACERO

df_peso= df_REDI.groupby(['Proyecto','Categoría'])[PESO].sum().reset_index()
df_peso.rename(columns={PESO: 'Peso(Kg)'}, inplace=True)
df_peso['Peso(Tn)'] = df_peso['Peso(Kg)']/1000

df_soldadura= df_REDI[df_REDI['Desc.Corta'].str.startswith('SOLDADURA', na=False)]
df_soldadura= df_soldadura.groupby(['Proyecto','Categoría'])[CTD].sum().reset_index()
df_soldadura.rename(columns={CTD: 'Soldadura(kg)'}, inplace=True)

df_alambre= df_REDI[df_REDI['Desc.Corta'].str.startswith('ALAMBRE', na=False)]
df_alambre= df_alambre.groupby(['Proyecto','Categoría'])[CTD].sum().reset_index()
df_alambre.rename(columns={CTD: 'Alambre tub(kg)'}, inplace=True)

df_oxigeno= df_REDI[df_REDI['Desc.Corta'].isin(['OXIGENO IND.'])]
df_oxigeno= df_oxigeno.groupby(['Proyecto','Categoría'])[CTD].sum().reset_index()
df_oxigeno.rename(columns={CTD: 'Oxigeno(m3)'}, inplace=True)

df_disco= df_REDI[df_REDI['Desc.Corta'].str.startswith('DISCO', na=False)]
df_disco= df_disco.groupby(['Proyecto','Categoría'])[CTD].sum().reset_index()
df_disco.rename(columns={CTD: 'Discos(pz)'}, inplace=True)

df_ratio = pd.merge(df_peso, df_soldadura, on=['Proyecto','Categoría'], how='outer')
df_ratio = pd.merge(df_ratio, df_alambre, on=['Proyecto','Categoría'], how='outer')
df_ratio = pd.merge(df_ratio, df_oxigeno, on=['Proyecto','Categoría'], how='outer')
df_ratio = pd.merge(df_ratio, df_disco, on=['Proyecto','Categoría'], how='outer')
df_ratio.fillna(0, inplace=True)

df_ratio['Soldadura Total(kg)'] = (df_ratio['Soldadura(kg)'] + df_ratio['Alambre tub(kg)']*1.67)

df_ratio['SoldxAcero'] = (df_ratio['Soldadura Total(kg)'])/df_ratio['Peso(Tn)']
df_ratio['OxigenoxAcero'] = df_ratio['Oxigeno(m3)']/df_ratio['Peso(Tn)']
df_ratio['DiscoxAcero'] = df_ratio['Discos(pz)']/df_ratio['Peso(Tn)']
df_ratio.fillna(0, inplace=True)

#############################################################################################################################
if selector_temporada == "Modulo":
    df_ratio_GRAFICO = df_ratio[df_ratio['Categoría'].isin(['VELA I CARENA','VELA I MODULO','SAMA I CARENA','SAMA I MODULO'])]
    color_domain_acero = ['VELA I CARENA','VELA I MODULO','SAMA I CARENA','SAMA I MODULO']
    color_range_acero = ['#2e86c1','#5dade2','#d6eaf8','#2e86c1']
    color_domain_sold = ['VELA I CARENA','VELA I MODULO','SAMA I CARENA','SAMA I MODULO']
    color_range_sold = ['#FFC300','#FF5733','#C70039','#FFC300']
    color_domain_oxi = ['VELA I CARENA','VELA I MODULO','SAMA I CARENA','SAMA I MODULO']
    color_range_oxi = ['#82e0aa ','#abebc6','#d5f5e3','#82e0aa ']
    color_domain_disc = ['VELA I CARENA','VELA I MODULO','SAMA I CARENA','SAMA I MODULO']
    color_range_disc = ['#85929e','#d6dbdf','#aeb6bf','#85929e']
else:
    df_ratio_GRAFICO = df_ratio[df_ratio['Categoría'].isin(['CASCO', 'ADITAMENTO', 'PANGA','PROYECTO MEJORA','OPEX','BE CASCO'])]   #SE FILTRA PARA CUADROS INICIALES DE RATIOS
    color_domain_acero = ['ADITAMENTO', 'CASCO', 'PANGA','PROYECTO MEJORA','OPEX','BE CASCO',]
    color_range_acero = ['#2e86c1','#5dade2','#d6eaf8','#2e86c1','#d6eaf8','#5dade2']
    color_domain_sold = ['ADITAMENTO', 'CASCO', 'PANGA','PROYECTO MEJORA','OPEX','BE CASCO',]
    color_range_sold = ['#FFC300','#FF5733','#C70039','#FFC300','#C70039','#FF5733']
    color_domain_oxi = ['ADITAMENTO', 'CASCO', 'PANGA','PROYECTO MEJORA','OPEX','BE CASCO']
    color_range_oxi = ['#82e0aa ','#abebc6','#d5f5e3','#82e0aa ','#d5f5e3','#abebc6']
    color_domain_disc = ['ADITAMENTO', 'CASCO', 'PANGA','PROYECTO MEJORA','OPEX','BE CASCO']
    color_range_disc = ['#85929e','#d6dbdf','#aeb6bf','#85929e','#aeb6bf','#d6dbdf']
#print(df_ratio)
col1, col2 =st.columns([0.5,0.5])
with col1:
    with st.container(border=True):
        st.subheader("Peso(Tn)")
        chart_ACERO = alt.Chart(df_ratio_GRAFICO[df_ratio_GRAFICO['Proyecto'].isin(selector_proyecto)]).mark_bar().encode(
            x=alt.X('Proyecto:N', title='', axis=alt.Axis(labelAngle=-35)),  # Título del eje X oculto
            y=alt.Y("Peso(Tn):Q", title='Peso(Tn)', axis=None),  # El dominio comienza en 0
            color=alt.Color('Categoría:N', scale=alt.Scale(domain=color_domain_acero, range=color_range_acero), legend=alt.Legend(title='Categoría',orient='bottom')),
            tooltip=['Proyecto', 'Categoría', 'Peso(Tn)']
        ).properties(width=300, height=350)

        st.altair_chart(chart_ACERO, use_container_width=True)
with col2:
    with st.container(border=True):
        st.subheader("Soldadura(Kg) vs Peso(Tn)")
        chart_Sold = alt.Chart(df_ratio_GRAFICO[df_ratio_GRAFICO['Proyecto'].isin(selector_proyecto)]).mark_bar().encode(
            x=alt.X('Proyecto:N', title='', axis=alt.Axis(labelAngle=-35)),  # Título del eje X oculto
            y=alt.Y("SoldxAcero:Q", title='SoldxAcero Kg/Tn', axis=None),  # El dominio comienza en 0
            color=alt.Color('Categoría:N', scale=alt.Scale(domain=color_domain_sold, range=color_range_sold), legend=alt.Legend(title='Categoría',orient='bottom')),
           tooltip=['Proyecto', 'Categoría', 'SoldxAcero']
        ).properties(width=300, height=350)

        st.altair_chart(chart_Sold, use_container_width=True)
        
col1, col2 =st.columns([0.5,0.5])
with col1:
    with st.container(border=True):
        st.subheader("Oxígeno(m3) vs Peso(Tn)")
        chart_oxigeno = alt.Chart(df_ratio_GRAFICO[df_ratio_GRAFICO['Proyecto'].isin(selector_proyecto)]).mark_bar().encode(
            x=alt.X('Proyecto:N', title='', axis=alt.Axis(labelAngle=-35)),  # Título del eje X oculto
            y=alt.Y("OxigenoxAcero:Q", title='OxigenoxAcero m3/Tn', axis=None),  # El dominio comienza en 0
            color=alt.Color('Categoría:N', scale=alt.Scale(domain=color_domain_oxi, range=color_range_oxi), legend=alt.Legend(title='Categoría',orient='bottom')),
            tooltip=['Proyecto', 'Categoría', 'OxigenoxAcero']
        ).properties(width=300, height=350)

        st.altair_chart(chart_oxigeno, use_container_width=True)
with col2:
    with st.container(border=True):
        st.subheader("Discos(pz) vs Peso(Tn)")
        chart_Sold = alt.Chart(df_ratio_GRAFICO[df_ratio_GRAFICO['Proyecto'].isin(selector_proyecto)]).mark_bar().encode(
            x=alt.X('Proyecto:N', title='', axis=alt.Axis(labelAngle=-35)),  # Título del eje X oculto
            y=alt.Y("DiscoxAcero:Q", title='Discos pz/Tn', axis=None),  # El dominio comienza en 0
            color=alt.Color('Categoría:N', scale=alt.Scale(domain=color_domain_disc, range=color_range_disc), legend=alt.Legend(title='Categoría',orient='bottom')),
           tooltip=['Proyecto', 'Categoría', 'DiscoxAcero']
        ).properties(width=300, height=350)

        st.altair_chart(chart_Sold, use_container_width=True)
############################################################################################################     
# Verifica si hay proyectos seleccionados
if selector_proyecto:
    # Filtrar los datos para los proyectos seleccionados
    df_general = df_UTI[df_UTI['Proyecto'].isin(selector_proyecto)]
    df_grouped = df_general.groupby(['Proyecto', 'Categoría']).agg({'MOD': 'sum', MAT: 'sum'}).reset_index()
    df_grouped['Total'] = (df_grouped['MOD'] + df_grouped[MAT]).round(2)
    
    # Crear listas dinámicas para las categorías y proyectos
    projects = df_grouped["Proyecto"].unique().tolist()  # Proyectos en el eje Y
    categories = df_grouped["Categoría"].unique().tolist()  # Categorías como series

    # Crear las series dinámicamente
    series = []
    for col in categories:
        # Obtener los valores para cada categoría en todos los proyectos
        category_data = df_grouped[df_grouped['Categoría'] == col]
        data = []
        for project in projects:
            total_value = category_data[category_data['Proyecto'] == project]['Total'].sum()
            data.append(total_value)
        
        # Añadir la serie con los datos por categoría
        series.append({
            "name": col,
            "type": "bar",
            "stack": "total",
            "label": {
                "show": True,
                "formatter": "S/. {c}",  # Formato para las etiquetas
            },
            "emphasis": {"focus": "series"},
            "data": data,
        })

    # Configurar las opciones del gráfico
    
    options = {
        "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
        "legend": {"data": categories},  # Nombres de las categorías como leyenda
        "grid": {"left": "3%", "right": "4%", "bottom": "3%", "containLabel": True},
        "xAxis": {"type": "value"},
        "yAxis": {"type": "category", "data": projects},  # Proyectos en el eje Y
        "series": series,
    }

    # Mostrar el gráfico en Streamlit
    with st.container(border=True):
            st.subheader("Proyectos")
            st_echarts(options=options, height="500px")
############################################################################################################     
# Verifica si hay proyectos seleccionados
if selector_proyecto:
    df_precioxpl = df_REDI[df_REDI['Proyecto'].isin(selector_proyecto)]
    df_precioxpl = df_precioxpl[['Proyecto','Desc.Corta',CTD,MAT]]
    df_filtrado = df_precioxpl[df_precioxpl['Desc.Corta'].isin(['PLANCHA AC.NAVAL 9.50X1.800X6000MM', 'PLANCHA AC.NAVAL 8.00X1.800X6000MM', 'PLANCHA AC.NAVAL 6.40X1.800X6000MM'])]
    df_agrupado = df_filtrado.groupby(['Proyecto', 'Desc.Corta'], as_index=False).agg({
        CTD: 'sum',
        MAT: 'sum'
    })
    
    # Calcular precio unitario
    df_agrupado['Precio Unitario'] = df_agrupado[MAT] / df_agrupado[CTD]
    
    # Preparar datos para el gráfico
    proyectos = df_agrupado['Proyecto'].unique().tolist()
    tipos_plancha = df_agrupado['Desc.Corta'].unique().tolist()
    
    # Crear series para cada tipo de plancha
    series = []
    for tipo in tipos_plancha:
        datos = []
        for proyecto in proyectos:
            valor = df_agrupado[(df_agrupado['Proyecto'] == proyecto) & 
                               (df_agrupado['Desc.Corta'] == tipo)]['Precio Unitario']
            datos.append(round(valor.values[0], 2) if len(valor) > 0 else 0)
        
        # Nombre corto para la leyenda
        nombre_corto = tipo.replace('PLANCHA AC.NAVAL ', 'PL-').replace('X1.800X6000MM', '')
        
        series.append({
            "name": nombre_corto,
            "type": "bar",
            "data": datos,
            "label": {"show": True, "position": "top", "formatter": "S/ {c}"}
        })
    
    # Configurar opciones del gráfico
    colores = ['#FFC107', '#4CAF50', '#2196F3']  # Amarillo, Verde, Azul
    
    options_planchas = {
        #"title": {"text": "Precio Unitario de Planchas"},
        "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
        "legend": {
            "data": [s["name"] for s in series],
            "top": "0%",  # Posición arriba
            "orient": "horizontal"  # Orientación horizontal
        },
        "grid": {"left": "3%", "right": "4%", "bottom": "3%", "containLabel": True},
        "xAxis": {"type": "category", "data": proyectos},
        "yAxis": {"type": "value", "name": "Precio Unitario (S/)"},
        "color": colores,
        "series": series,
    }
    
    # Mostrar el gráfico
    with st.container(border=True):
        st.subheader("Precio Unitario de Plancha")
        st_echarts(options=options_planchas, height="400px")



############################################################################################################   
if selector_proyecto:
 for proyecto in selector_proyecto:
    with st.container(border=True):    
        df_peso = df_peso[df_peso['Proyecto'] == proyecto]
        df_soldadura = df_soldadura[df_soldadura['Proyecto'] == proyecto]
        df_alambre = df_alambre[df_alambre['Proyecto'] == proyecto]   
        df_oxigeno = df_oxigeno[df_oxigeno['Proyecto'] == proyecto] 
        
        st.subheader(proyecto, divider=True)
        # Filtrar datos del proyecto seleccionado
        df_general = df_UTI[df_UTI['Proyecto'] == proyecto]
        df1= df_general.groupby(['Proyecto'])['MOD'].sum().reset_index()
        df2= df_general.groupby(['Proyecto'])[MAT].sum().reset_index()
        result = pd.merge(df1, df2, on='Proyecto')
        result['Total'] = result['MOD'] + result[MAT]
        result['Total USD'] = result['Total'] / precio_dolar
        
        # Dar formato de moneda a los datos
        formatted_result = result.style.format({'MOD': 'S/. {:,.2f}', MAT: 'S/. {:,.2f}', 'Total': 'S/. {:,.2f}', 'Total USD': '$. {:,.2f}'})

        # Mostrar el DataFrame en Streamlit
        st.dataframe(formatted_result,hide_index=True, use_container_width=True)

        #Costos x PEP
        df_PEP = df_general['Categoría'].dropna().drop_duplicates().sort_values()
        df_PEP2 = df_general.groupby(['Categoría'])['MOD'].sum().reset_index()
        df_PEP3 = df_general.groupby(['Categoría'])[MAT].sum().reset_index()
        result2 = pd.merge(df_PEP2, df_PEP3, on='Categoría')
        result2['Total'] = result2['MOD'] + result2[MAT]
        result2['Total USD'] = result2['Total'] / precio_dolar

        # Dar formato de moneda a los datos
        formatted_result = result2.style.format({'MOD': 'S/. {:,.2f}', MAT: 'S/. {:,.2f}', 'Total': 'S/. {:,.2f}', 'Total USD': '$. {:,.2f}'})
        st.dataframe(formatted_result,hide_index=True, use_container_width=True)
        

        df_ratio_proyecto = df_ratio[df_ratio['Proyecto'] == proyecto]


        st.subheader("Acero")
        col1, col2 =st.columns([0.7,0.3])
        with col1:
            st.dataframe(df_ratio_proyecto, use_container_width=False,column_config={
                        "Proyecto": None,
                        "Soldadura(kg)": None,
                        "Alambre tub(kg)": None,
                        "Peso(Tn)": None,
                        "OxigenoxAcero ": None,
                        "DiscoxAcero ": None,
                        'Peso(Kg)': st.column_config.NumberColumn(
                            "Peso(Kg)",
                            format="%.2f Kg",
                            width=None,
                        )

                    },hide_index=True)
       
        with col2:
            chart = alt.Chart(df_ratio_proyecto).mark_arc(innerRadius=50).encode(
            theta=alt.Theta(field="Peso(Kg)", type="quantitative", title="Peso(kg)"),
            color=alt.Color(field="Categoría", type="nominal", legend=alt.Legend(title="Categorías")),
            tooltip=["Categoría", alt.Tooltip('Peso(Kg)', title="Peso(kg)")]
            ).properties(
                width=300,
                height=300
            )
            
            # Mostrar el gráfico
            st.altair_chart(chart, use_container_width=True)

              
else:
    st.info("Por favor, seleccione uno o más proyectos para ver los detalles.")
    


