import streamlit as st
import pandas as pd
import requests
import numpy as np
import altair as alt
#from streamlit_lottie import st_lottie
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook

# Configuración de la página
st.set_page_config(
    page_title="Control de Proyectos",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded")


# Obtener la fecha actual en el formato deseado
FECHA = datetime.now().strftime('%d-%m-%y')

# Ruta al archivo de Excel
BD = './BD.xlsx'

# Cargar el libro de trabajo y las hojas disponibles
TEMPORADAS = load_workbook(BD, read_only=True).sheetnames

# Configuración de la barra lateral
st.sidebar.header('Proyectos :anchor:')
selector_temporada = st.sidebar.selectbox("Seleccione la temporada:", TEMPORADAS, index=0)
df_proyecto = pd.read_excel(BD, sheet_name=selector_temporada)

# Mostrar opciones para seleccionar proyectos
#selector_proyecto = st.sidebar.multiselect("Seleccione proyectos:", df_proyecto['Proyecto'].drop_duplicates())

# Ruta al archivo de Excel
UTI = './'+selector_temporada+'/UTI.xlsx'
df_UTI = pd.read_excel(UTI, sheet_name="Sheet1")

# Reemplazar valores N/A (NaN) con ceros
df_UTI = df_UTI.fillna(0)

REDI = './'+selector_temporada+'/REDI.xlsx'
df_REDI = pd.read_excel(REDI, sheet_name="Sheet1")


select_proyecto = st.sidebar.selectbox('Seleciona un proyecto',df_UTI['Proyecto'].dropna().drop_duplicates().sort_values(), key='elect_proyecto')

PD_BD = df_UTI[df_UTI['Proyecto'].isin([select_proyecto])]

select_sistema = st.sidebar.selectbox('Seleciona un sistema',PD_BD['Categoría'].dropna().drop_duplicates().sort_values(), key='select_sistema')

PD_BD = PD_BD[PD_BD['Categoría'].isin([select_sistema])]

#Ocultar para consolidado de pesos
#select_actividad = st.sidebar.selectbox('Seleciona una actividad',PD_MAT['Descripción Grafo'].dropna().drop_duplicates().sort_values(), key='actividad')

#PD_MAT = PD_MAT[PD_MAT['Descripción Grafo'].isin([select_actividad])]
###################################################################

PD_MAT = df_REDI.drop(columns=['Tratar','Ind.REDI','Nro.REDI'])
PD_MAT = PD_MAT[PD_MAT['Proyecto'].isin([select_proyecto])]
PD_MAT = PD_MAT[PD_MAT['Categoría'].isin([select_sistema])]

st.title (select_proyecto)

tab1, tab2 = st.tabs(["Actividades", "Materiales"])

with tab1:

    st.dataframe(PD_BD,hide_index=True, use_container_width=True, height=600, column_config={
        'Grafo': st.column_config.NumberColumn(
                "Grafo",
                format="%d",
                width=None
            ),     
        'Oper.': st.column_config.NumberColumn(
                "Oper.",
                width="small"
            ),  
        'MOD': st.column_config.NumberColumn(
                "MOD",
                format="S/.%.2f",          
            ),
        'MAT Despachado': st.column_config.NumberColumn(
                "MAT Despachado",
                format="S/.%.2f",
            ),
        'MAT Estimado': st.column_config.NumberColumn(
                "MAT Estimado",
                format="S/.%.2f",
            )      
        },column_order=("Grafo", "Oper.","Descripción Grafo","Denom.Operación", "MOD", "MAT Estimado","MAT Despachado","Nombre Acreedor"))

with tab2:
    st.dataframe(PD_MAT,hide_index=True, use_container_width=True, height=600, column_config={
    "Descripción Grafo":None,
    "Proyecto":None,
    "Categoría":None,
    "Oper.":None,
    "Material":None,
    "Reserva":None,
    "Fe.Necesidad":None,
    'Grafo': st.column_config.NumberColumn(
            "Grafo",
            format="%d",
        ),       
    'MAT Estimado': st.column_config.NumberColumn(
            "MAT Est.",
            format="S/.%.3f",
        ),
    'MAT Despachado': st.column_config.NumberColumn(
            "MAT tomado",
            format="S/.%.3f",
        ) ,
    "Peso(kg)": st.column_config.NumberColumn(
            "Peso(kg)",
            format="%.2f kg",
            width="None",
        )   ,
    "Peso estimado(kg)": st.column_config.NumberColumn(
            "Peso est.(kg)",
            format="%.2f kg",
            width="None",
        )
    },column_order=("Grafo", "Denom.Operación","Desc.Corta","UMB","Cantidad","Cantidad tomada", "MAT Estimado","MAT Despachado","Ubicación","Dim1","Dim2","Dim3","Peso(kg)","Peso estimado(kg)"))

    #################################################################
    #Cantidades tomadas
    df1= PD_MAT.groupby(['Denom.Operación'])['Peso(kg)'].sum().reset_index()
    df2= PD_MAT[PD_MAT['Desc.Corta'].isin(['OXIGENO IND.'])]
    df3= PD_MAT[PD_MAT['Desc.Corta'].str.contains('disco', case=False, na=False)]
    df4= PD_MAT[PD_MAT['Desc.Corta'].str.contains('PROPANO', case=False, na=False)]
    df5= PD_MAT[PD_MAT['Desc.Corta'].str.startswith('SOLDADURA', na=False)]
    df6= PD_MAT[PD_MAT['Desc.Corta'].str.startswith('ALAMBRE', na=False)]

    df2= df2.groupby(['Denom.Operación'])['Cantidad tomada'].sum().reset_index()
    df2.rename(columns={'Cantidad tomada': 'Oxígeno(m3)'}, inplace=True)

    df3= df3.groupby(['Denom.Operación'])['Cantidad tomada'].sum().reset_index()
    df3.rename(columns={'Cantidad tomada': 'Discos'}, inplace=True)

    df4= df4.groupby(['Denom.Operación'])['Cantidad tomada'].sum().reset_index()
    df4.rename(columns={'Cantidad tomada': 'Gas Propano(bot)'}, inplace=True)

    df5= df5.groupby(['Denom.Operación'])['Cantidad tomada'].sum().reset_index()
    df5.rename(columns={'Cantidad tomada': 'Soldadura(kg)'}, inplace=True)

    df6= df6.groupby(['Denom.Operación'])['Cantidad tomada'].sum().reset_index()
    df6.rename(columns={'Cantidad tomada': 'Alambre tub(kg)'}, inplace=True)

    #Cantidad reservada
    dfA= PD_MAT.groupby(['Denom.Operación'])['Peso estimado(kg)'].sum().reset_index()
    dfB= PD_MAT[PD_MAT['Desc.Corta'].isin(['OXIGENO IND.'])]
    dfC= PD_MAT[PD_MAT['Desc.Corta'].str.contains('disco', case=False, na=False)]
    dfD= PD_MAT[PD_MAT['Desc.Corta'].str.contains('PROPANO', case=False, na=False)]
    dfE= PD_MAT[PD_MAT['Desc.Corta'].str.startswith('SOLDADURA', na=False)]
    dfF= PD_MAT[PD_MAT['Desc.Corta'].str.startswith('ALAMBRE', na=False)]

    dfA.rename(columns={'Peso estimado(kg)': 'Peso(kg)'}, inplace=True)

    dfB= dfB.groupby(['Denom.Operación'])['Cantidad'].sum().reset_index()
    dfB.rename(columns={'Cantidad': 'Oxígeno(m3)'}, inplace=True)

    dfC= dfC.groupby(['Denom.Operación'])['Cantidad'].sum().reset_index()
    dfC.rename(columns={'Cantidad': 'Discos'}, inplace=True)

    dfD= dfD.groupby(['Denom.Operación'])['Cantidad'].sum().reset_index()
    dfD.rename(columns={'Cantidad': 'Gas Propano(bot)'}, inplace=True)

    dfE= dfE.groupby(['Denom.Operación'])['Cantidad'].sum().reset_index()
    dfE.rename(columns={'Cantidad': 'Soldadura(kg)'}, inplace=True)

    dfF= dfF.groupby(['Denom.Operación'])['Cantidad'].sum().reset_index()
    dfF.rename(columns={'Cantidad': 'Alambre tub(kg)'}, inplace=True)

    # Unir DATA TOMADA
    df_tomada = pd.merge(df1, df2, on='Denom.Operación', how='outer')
    df_tomada = pd.merge(df_tomada, df3, on='Denom.Operación', how='outer')
    df_tomada = pd.merge(df_tomada, df4, on='Denom.Operación', how='outer')
    df_tomada = pd.merge(df_tomada, df5, on='Denom.Operación', how='outer')
    df_tomada = pd.merge(df_tomada, df6, on='Denom.Operación', how='outer')

    # Unir DATA ESTIMADA
    df_estimada = pd.merge(dfA, dfB, on='Denom.Operación', how='outer')
    df_estimada = pd.merge(df_estimada, dfC, on='Denom.Operación', how='outer')
    df_estimada = pd.merge(df_estimada, dfD, on='Denom.Operación', how='outer')
    df_estimada = pd.merge(df_estimada, dfE, on='Denom.Operación', how='outer')
    df_estimada = pd.merge(df_estimada, dfF, on='Denom.Operación', how='outer')


    st.subheader("Cantidad tomada")
    st.dataframe(df_tomada,column_config={
            "Denom.Operación": st.column_config.Column(
                "Denom.Operación",
                width="medium",
            ),
            "Peso(kg)": st.column_config.NumberColumn(
                "Peso(kg)",
                format="%.2f kg",
                width="None",
            )
        },hide_index=True)

    st.subheader("Cantidad REDI")
    st.dataframe(df_estimada,column_config={
            "Denom.Operación": st.column_config.Column(
                "Denom.Operación",
                width="medium",
            ),
            "Peso(kg)": st.column_config.NumberColumn(
                "Peso(kg)",
                format="%.2f kg",
                width="None",
            )
        },hide_index=True)
