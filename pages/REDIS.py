import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import altair as alt
import numpy as np
from streamlit_extras.colored_header import colored_header
import matplotlib.pyplot as plt
# Fecha actual
FECHA = datetime.now().strftime('%d-%m-%y')

# Configuración de la página
st.set_page_config(
    page_title="Control de Proyectos",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded")

# Ruta al archivo de Excel
BD = './BD.xlsx'

# Cargar el libro de trabajo y las hojas disponibles
TEMPORADAS = load_workbook(BD, read_only=True).sheetnames

# Configuración de la barra lateral
st.sidebar.header('Proyectos :anchor:')
selector_temporada = st.sidebar.selectbox("Seleccione la temporada:", TEMPORADAS, index=0)
df_proyecto = pd.read_excel(BD, sheet_name=selector_temporada)

# Mostrar opciones para seleccionar proyectos
selector_proyecto = st.sidebar.selectbox("Seleccione proyectos:", df_proyecto['Proyecto'].drop_duplicates())

# Título de la aplicación
st.title(":ship: "+selector_proyecto)

# Ruta al archivo de Excel
UTI = './'+selector_temporada+'/UTI.xlsx'
df_UTI = pd.read_excel(UTI, sheet_name="Sheet1")

# Reemplazar valores N/A (NaN) con ceros
df_UTI = df_UTI.fillna(0)

REDI = './'+selector_temporada+'/REDI.xlsx'
df_REDI = pd.read_excel(REDI, sheet_name="Sheet1")

df_UTI = df_UTI[df_UTI['Proyecto'].isin([selector_proyecto])]

selector_proveedor = st.sidebar.selectbox("Seleccione proveedor:", df_UTI['Nombre Acreedor'].drop_duplicates().replace(0, pd.NA).dropna())

df_UTI = df_UTI[df_UTI['Nombre Acreedor'].isin([selector_proveedor])]

df_REDI = df_REDI[df_REDI['Proyecto'].isin([selector_proyecto])]

selector_grafo = st.sidebar.multiselect("Seleccione:", df_UTI['Descripción Grafo'].drop_duplicates().replace(0, pd.NA).dropna())

df_UTI = df_UTI[df_UTI['Descripción Grafo'].isin(selector_grafo)]

keys = df_UTI.set_index(['Grafo', 'Oper.']).index  #Filtrar por Grafo y Operación
df_REDI = df_REDI[df_REDI.set_index(['Grafo', 'Oper.']).index.isin(keys)].reset_index(drop=True)

selector_actividad = df_REDI["Denom.Operación"].drop_duplicates()

#SUMA DE ACERO
df_acero= df_REDI.groupby(['Proyecto','Categoría'])['Peso(kg)'].sum().reset_index()
df_acero["Peso(kg)"] = df_acero["Peso(kg)"]/1000
df_acero.rename(columns={'Peso(kg)': 'Peso(Tn)'}, inplace=True)

df_soldadura= df_REDI[df_REDI['Desc.Corta'].str.startswith('SOLDADURA', na=False)]
df_soldadura= df_soldadura.groupby(['Proyecto','Categoría'])['Cantidad tomada'].sum().reset_index()
df_soldadura.rename(columns={'Cantidad tomada': 'Soldadura(kg)'}, inplace=True)

df_alambre= df_REDI[df_REDI['Desc.Corta'].str.startswith('ALAMBRE', na=False)]
df_alambre= df_alambre.groupby(['Proyecto','Categoría'])['Cantidad tomada'].sum().reset_index()
df_alambre.rename(columns={'Cantidad tomada': 'Alambre tub(kg)'}, inplace=True)

df_oxigeno= df_REDI[df_REDI['Desc.Corta'].isin(['OXIGENO IND.'])]
df_oxigeno= df_oxigeno.groupby(['Proyecto','Categoría'])['Cantidad tomada'].sum().reset_index()
df_oxigeno.rename(columns={'Cantidad tomada': 'Oxigeno(m3)'}, inplace=True)

df_disco= df_REDI[df_REDI['Desc.Corta'].str.startswith('DISCO', na=False)]
df_disco= df_disco.groupby(['Proyecto','Categoría'])['Cantidad tomada'].sum().reset_index()
df_disco.rename(columns={'Cantidad tomada': 'Discos(pz)'}, inplace=True)

df_ratio_acero = pd.merge(df_acero, df_soldadura, on=['Proyecto','Categoría'], how='outer')
df_ratio_acero = pd.merge(df_ratio_acero, df_alambre, on=['Proyecto','Categoría'], how='outer')
df_ratio_acero = pd.merge(df_ratio_acero, df_oxigeno, on=['Proyecto','Categoría'], how='outer')
df_ratio_acero = pd.merge(df_ratio_acero, df_disco, on=['Proyecto','Categoría'], how='outer')
df_ratio_acero.fillna(0, inplace=True)

#############################################################################################################################
    
# Verifica si hay proyectos seleccionados
if selector_proveedor:
 for actividad in selector_actividad:
    df_REDI_filtrado = df_REDI[df_REDI['Denom.Operación'].isin([actividad])]
    st.subheader(actividad, divider=True)
    # Filtrar datos del proyecto seleccionado
    #print(df_REDI_filtrado)

    st.dataframe(df_REDI_filtrado,use_container_width=False,column_config={
                        "Tratar": None,
                        "Proyecto": None,
                        "Denom.Operación": None,
                        "Categoría": None,
                        "Grafo": None,
                        "Oper.": None,
                        "MAT Despachado": None,
                        "MAT Estimado": None,
                        "Reserva": None,
                        "Material": None,
                        "Fe.Necesidad": None,
                        "Ind.REDI": None,
                        "Nro.REDI": None,
                        'Peso(kg)': st.column_config.NumberColumn(
                            "Peso(Kg)",
                            format="%.2f Kg",
                            width=None,
                        ),
                        "Desc.Corta": st.column_config.Column(width='large')

                    },hide_index=True)
                
else:
    st.info("Por favor, seleccione uno o más proyectos para ver los detalles.")
    

