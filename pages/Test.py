import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os
from PIL import Image
from datetime import datetime, timedelta
import plotly.express as px
import json

# Fecha actual
FECHA = datetime.now().strftime('%d-%m-%y')

def obtener_rango_semana(fecha):
    # Asegurarse de que sea un objeto de tipo fecha
    if not isinstance(fecha, datetime):
        fecha = pd.to_datetime(fecha)  # Convertir a datetime si es necesario
    
    # Calcular el lunes y el domingo de la semana
    lunes = fecha - timedelta(days=fecha.weekday())
    domingo = lunes + timedelta(days=6)
    
    return lunes.strftime("%d/%m/%Y"), domingo.strftime("%d/%m/%Y")

# Configuración de la página
st.set_page_config(
    page_title="Control de Proyectos",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Ruta al archivo de Excel
BD = './BD.xlsx'

# Leer todas las hojas del archivo de Excel
hojas = pd.read_excel(BD, sheet_name=None)  # `sheet_name=None` carga todas las hojas

# Crear una lista para almacenar los DataFrames de todas las hojas
dataframes = []

# Iterar sobre cada hoja y agregarla a la lista de DataFrames
for nombre_hoja, df in hojas.items():
    df['Hoja'] = nombre_hoja
    dataframes.append(df)

# Concatenar todos los DataFrames de la lista en uno solo
df_total = pd.concat(dataframes, ignore_index=True)

# Cargar el libro de trabajo y las hojas disponibles
TEMPORADAS = load_workbook(BD, read_only=True).sheetnames

# Selección de temporada
selector_temporada = st.sidebar.selectbox("Seleccione la temporada:", TEMPORADAS)

# Filtrar los datos por la temporada seleccionada
df_proyecto = df_total[df_total['Temporada'].isin([selector_temporada])]

# Selección de proyecto
selector_proyecto = st.sidebar.selectbox("Seleccione proyecto:", df_proyecto['Proyecto'].drop_duplicates().sort_values(), index=0)

df_BD= df_proyecto[df_proyecto['Proyecto'].isin([selector_proyecto])]
Inicio_proyecto =df_BD["Inicio"].iloc[0]

# Reemplazar "/" en el nombre del proyecto para evitar problemas con rutas
selector_proyecto_formateado = selector_proyecto.replace("/", "")

# Ruta de la carpeta del proyecto
carpeta_proyecto = os.path.join("./Bitacora/"+selector_temporada, selector_proyecto_formateado)

st.subheader("🚢 " + selector_proyecto_formateado)

col1, col2 = st.columns(2)

with col1:
    # Ruta al archivo JSON externo
    json_path = "proyectos.json"

    # Cargar el JSON
    with open(json_path, "r") as file:
        proyectos_json = json.load(file)
        
    # Convertir el JSON en un formato manejable
    data = proyectos_json["Proyectos"]

    try:
        # Intentar obtener el proyecto del JSON
        proyecto = next(p for p in data if p["Nombre"] == selector_proyecto_formateado)
        
        # Acceder a las claves de forma segura usando get()
        fechas = proyecto.get("Fechas", [])
        avance_programado = proyecto.get("AvanceProgramado", [])
        avance_real = proyecto.get("AvanceReal", [])
        
        # Verificar si las listas de fechas y avances están vacías
        if not fechas or not avance_programado or not avance_real:
            st.warning("Algunos datos del proyecto están incompletos (Fechas o Avances).")
    
    except StopIteration:
        st.warning(f"No se encontró el proyecto '{selector_proyecto_formateado}' en los datos.")
        proyecto = None
    except KeyError:
        st.error("La estructura del archivo JSON no es la esperada. No se encuentra la clave 'Nombre'.")
        proyecto = None

    # Verificar si el proyecto tiene datos completos
    if proyecto and fechas and avance_programado and avance_real:
        # Crear el DataFrame con los datos del proyecto
        df = pd.DataFrame({
            "Fechas": fechas,
            "Avance Programado": avance_programado,
            "Avance Real": avance_real
        })

        # Crear el gráfico con Plotly
        fig = px.line(
            df,
            x="Fechas",
            y=["Avance Programado", "Avance Real"],
            title=f"Curvas de Avance - {selector_proyecto_formateado}",
            labels={"value": "Porcentaje", "variable": "Tipo de Avance"},
            markers=True
        )

        # Personalizar el diseño del gráfico
        fig.update_layout(
            xaxis_title="Fechas",
            yaxis_title="Porcentaje (%)",
            legend_title="Tipo de Avance",
            template="plotly_white"
        )

        # Mostrar el gráfico
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.write("Datos insuficientes para generar el gráfico.")

with col2:
    with st.container(height=600):
        # Verificar si la carpeta del proyecto existe
        if os.path.exists(carpeta_proyecto):
            # Buscar todas las subcarpetas dentro de la carpeta del proyecto
            subcarpetas = [f for f in os.listdir(carpeta_proyecto) if os.path.isdir(os.path.join(carpeta_proyecto, f))]

            # Iterar sobre cada subcarpeta para crear un carrusel
            for subcarpeta in sorted(subcarpetas):
                try:
                    # Convertir el nombre de la subcarpeta a un número (por ejemplo, "1", "2", ...)
                    semana_numero = int(subcarpeta)
                except ValueError:
                    st.error(f"La subcarpeta '{subcarpeta}' no tiene un formato válido para semanas.")
                    continue

                # Calcular la fecha inicial de la semana actual
                if semana_numero == 1:
                    # Si es la primera semana, usar la fecha de inicio del proyecto directamente
                    fecha_semana_actual = Inicio_proyecto
                else:
                    # Para semanas posteriores, sumar semanas al inicio del proyecto
                    fecha_semana_actual = Inicio_proyecto + timedelta(weeks=semana_numero - 1)

                lunes, domingo = obtener_rango_semana(fecha_semana_actual)

                # Mostrar la subcabecera con el rango de la semana
                st.subheader(f"Semana {subcarpeta}: {lunes} - {domingo}", divider=True)

                carpeta_imagenes = os.path.join(carpeta_proyecto, subcarpeta)
                
                # Leer el archivo de descripción (si existe)
                descripcion_path = os.path.join(carpeta_imagenes, "descripcion.txt")
                descripcion = ""
                if os.path.exists(descripcion_path):
                    with open(descripcion_path, "r", encoding="utf-8") as f:
                        descripcion = f.read().strip()
                    descripcion = descripcion.replace("\n", "  \n")  # Formato Markdown para salto de línea
                    with st.expander("Ver detalle"):
                        st.markdown(descripcion)
                else:
                    st.info("No se encontró una descripción para esta carpeta.")

                # Verificar si hay imágenes en la subcarpeta
                imagenes = [os.path.join(carpeta_imagenes, img) for img in os.listdir(carpeta_imagenes) if img.endswith((".png", ".jpg", ".jpeg"))]

                if imagenes:
                    # Crear un índice inicial para el carrusel específico de esta subcarpeta
                    key_img_index = f"img_index_{subcarpeta}"
                    if key_img_index not in st.session_state:
                        st.session_state[key_img_index] = 0

                    # Mostrar la imagen actual
                    img_actual = Image.open(imagenes[st.session_state[key_img_index]])
                    st.image(img_actual, caption=f"Imagen {st.session_state[key_img_index] + 1} de {len(imagenes)}")

                    # Navegación del carrusel
                    col1, col2, col3, col4 = st.columns([0.5, 1, 2, 1])

                    with col2:
                        if st.button("⬅️ Anterior", key=f"prev_{subcarpeta}"):
                            st.session_state[key_img_index] = (st.session_state[key_img_index] - 1) % len(imagenes)

                    with col4:
                        if st.button("➡️ Siguiente", key=f"next_{subcarpeta}"):
                            st.session_state[key_img_index] = (st.session_state[key_img_index] + 1) % len(imagenes)
                else:
                    st.warning(f"No se encontraron imágenes en la carpeta {subcarpeta}.")
