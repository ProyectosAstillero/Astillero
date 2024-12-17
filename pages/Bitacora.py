import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os
from PIL import Image

# Fecha actual
FECHA = datetime.now().strftime('%d-%m-%y')

# Configuración de la página
st.set_page_config(
    page_title="Control de Proyectos",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Ruta al archivo de Excel
BD = './BD.xlsx'

# Leer todas las hojas del archivo de Excel
hojas = pd.read_excel(BD, sheet_name=None)  # `sheet_name=None` carga todas las hojas

# Crear una lista para almacenar los DataFrames de todas las hojas
dataframes = []

# Iterar sobre cada hoja y agregarla a la lista de DataFrames
for nombre_hoja, df in hojas.items():
    df['Hoja'] = nombre_hoja
    dataframes.append(df)

# Concatenar todos los DataFrames de la lista en uno solo
df_total = pd.concat(dataframes, ignore_index=True)

# Cargar el libro de trabajo y las hojas disponibles
TEMPORADAS = load_workbook(BD, read_only=True).sheetnames

# Selección de temporada
selector_temporada = st.sidebar.selectbox("Seleccione la temporada:", TEMPORADAS)

# Filtrar los datos por la temporada seleccionada
df_proyecto = df_total[df_total['Temporada'].isin([selector_temporada])]

# Selección de proyecto
selector_proyecto = st.sidebar.selectbox("Seleccione proyecto:", df_proyecto['Proyecto'].drop_duplicates().sort_values(), index=0)

# Reemplazar "/" en el nombre del proyecto para evitar problemas con rutas
selector_proyecto = selector_proyecto.replace("/", "")

# Ruta de la carpeta del proyecto
carpeta_proyecto = os.path.join(selector_temporada, selector_proyecto)

st.subheader("🚢 " + selector_proyecto)

# Verificar si la carpeta del proyecto existe
if os.path.exists(carpeta_proyecto):
    # Buscar todas las subcarpetas dentro de la carpeta del proyecto
    subcarpetas = [f for f in os.listdir(carpeta_proyecto) if os.path.isdir(os.path.join(carpeta_proyecto, f))]

    # Iterar sobre cada subcarpeta para crear un carrusel
    for subcarpeta in sorted(subcarpetas):
        st.subheader(f"Semana {subcarpeta}",divider=True)
        carpeta_imagenes = os.path.join(carpeta_proyecto, subcarpeta)
        
        # Leer el archivo de descripción (si existe)
        descripcion_path = os.path.join(carpeta_imagenes, "descripcion.txt")
        descripcion = ""
        if os.path.exists(descripcion_path):
            with open(descripcion_path, "r", encoding="utf-8") as f:
                 descripcion = f.read().strip()
            descripcion = descripcion.replace("\n", "  \n")  # Formato Markdown para salto de línea
            st.markdown(descripcion)

        else:
            st.info("No se encontró una descripción para esta carpeta.")
        # Verificar si hay imágenes en la subcarpeta
        imagenes = [os.path.join(carpeta_imagenes, img) for img in os.listdir(carpeta_imagenes) if img.endswith((".png", ".jpg", ".jpeg"))]

        if imagenes:
            # Crear un índice inicial para el carrusel específico de esta subcarpeta
            key_img_index = f"img_index_{subcarpeta}"
            if key_img_index not in st.session_state:
                st.session_state[key_img_index] = 0

            # Mostrar la imagen actual
            img_actual = Image.open(imagenes[st.session_state[key_img_index]])
            st.image(img_actual, caption=f"Imagen {st.session_state[key_img_index] + 1} de {len(imagenes)}")

            # Navegación del carrusel
            col1, col2, col3, col4= st.columns([0.5,1,2,1])

            with col2:
                if st.button("⬅️ Anterior", key=f"prev_{subcarpeta}"):
                    st.session_state[key_img_index] = (st.session_state[key_img_index] - 1) % len(imagenes)

            with col4:
                if st.button("➡️ Siguiente", key=f"next_{subcarpeta}"):
                    st.session_state[key_img_index] = (st.session_state[key_img_index] + 1) % len(imagenes)
        else:
            st.warning(f"No se encontraron imágenes en la carpeta {subcarpeta}.")
else:
    st.error(f"No existe la carpeta del proyecto: {carpeta_proyecto}.")
