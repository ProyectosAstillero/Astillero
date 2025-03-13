import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import altair as alt
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from streamlit_echarts import st_echarts

# Configuración de la página
st.set_page_config(
    page_title="Control de Proyectos",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Ruta al archivo de Excel
BD = './BD.xlsx'

# Leer todas las hojas del archivo de Excel
hojas = pd.read_excel(BD, sheet_name=None)  # `sheet_name=None` carga todas las hojas

# Crear una lista para almacenar los DataFrames de todas las hojas
dataframes = []

# Iterar sobre cada hoja y agregarla a la lista de DataFrames
for nombre_hoja, df in hojas.items():
    # Agregar una nueva columna 'Hoja' para identificar de qué hoja proviene cada fila
    df['Hoja'] = nombre_hoja
    # Agregar el DataFrame de la hoja a la lista
    dataframes.append(df)

# Concatenar todos los DataFrames de la lista en uno solo
df_total = pd.concat(dataframes, ignore_index=True)

# Cargar el libro de trabajo y las hojas disponibles
TEMPORADAS = load_workbook(BD, read_only=True).sheetnames

selector_temporada = st.sidebar.selectbox("Seleccione la temporada:", TEMPORADAS)

df_proyecto = df_total[df_total['Temporada'].isin({selector_temporada})]

print(selector_temporada)
# Inicializamos las listas para almacenar los DataFrames de cada temporada
df_UTI_list = []

# Construir la ruta al archivo UTI y REDI
UTI = './' + selector_temporada + '/UTI.xlsx'
    
# Cargar los archivos correspondientes a la temporada
df_UTI = pd.read_excel(UTI, sheet_name="Sheet1")
    
# Reemplazar valores N/A (NaN) con ceros en el DataFrame de UTI
df_UTI = df_UTI.fillna(0)
    
# Almacenar los DataFrames en las listas
df_UTI_list.append(df_UTI)

# Si necesitas combinar los DataFrames de todas las temporadas en uno solo, puedes usar pd.concat:
df_UTI = pd.concat(df_UTI_list, ignore_index=True)

PROVEEDORES = (
    df_UTI['Nombre Acreedor']
    .dropna()                           # Elimina NaN
    .astype(str)                        # Convierte todo a string
    .str.strip()                        # Elimina espacios en los extremos
    .loc[lambda x: (x != "") & (x != "0")]  # Filtra vacíos y "0" (usando .loc)
    .unique()                           # Obtiene valores únicos
    .tolist()                           # Convierte a lista
)
print(PROVEEDORES)
selector_proveedor = st.sidebar.multiselect("Seleccione proveedores:", PROVEEDORES)

df_filtrada = df_UTI[df_UTI['Nombre Acreedor'].isin(selector_proveedor)]

df_filtrada = df_filtrada[df_filtrada['Liquidación'] > 0]

# Crear un DataFrame con el total por proveedor
df_total_proveedor = df_filtrada.groupby('Nombre Acreedor')['MOD'].sum().reset_index()

# Crear un DataFrame con el detalle por proyecto y proveedor
df_detalle = df_filtrada.groupby(['Nombre Acreedor', 'Proyecto'])['MOD'].sum().reset_index()

# Crear gráfico de barras con Plotly
fig = go.Figure()

# Agregar barra para cada proveedor
fig.add_trace(go.Bar(
    x=df_total_proveedor['Nombre Acreedor'],
    y=df_total_proveedor['MOD'],
    name='Total MOD',
    text=df_total_proveedor['MOD'].round(2),
    textposition='auto',
))

# Configurar el diseño
fig.update_layout(
    title='Costos MOD por Proveedor',
    xaxis_title='Proveedor',
    yaxis_title='MOD (Soles)',
    barmode='group',
    height=500,
    showlegend=True
)

# Mostrar el gráfico de barras
st.plotly_chart(fig, use_container_width=True)

# Mostrar tabla de detalle por proyecto
st.subheader('Detalle de MOD por Proyecto y Proveedor')
df_detalle_formateado = df_detalle.copy()
df_detalle_formateado['MOD'] = df_detalle_formateado['MOD'].round(2)
st.dataframe(df_detalle_formateado, use_container_width=True, hide_index=True)