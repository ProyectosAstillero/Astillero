import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import altair as alt
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from streamlit_echarts import st_echarts
# Fecha actual
FECHA = datetime.now().strftime('%d-%m-%y')

# Configuración de la página
st.set_page_config(
    page_title="Control de Proyectos",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Ruta al archivo de Excel
BD = './BD.xlsx'

# Leer todas las hojas del archivo de Excel
hojas = pd.read_excel(BD, sheet_name=None)  # `sheet_name=None` carga todas las hojas

# Crear una lista para almacenar los DataFrames de todas las hojas
dataframes = []

# Iterar sobre cada hoja y agregarla a la lista de DataFrames
for nombre_hoja, df in hojas.items():
    # Agregar una nueva columna 'Hoja' para identificar de qué hoja proviene cada fila
    df['Hoja'] = nombre_hoja
    # Agregar el DataFrame de la hoja a la lista
    dataframes.append(df)

# Concatenar todos los DataFrames de la lista en uno solo
df_total = pd.concat(dataframes, ignore_index=True)

# Cargar el libro de trabajo y las hojas disponibles
TEMPORADAS = load_workbook(BD, read_only=True).sheetnames

selector_temporada = st.sidebar.multiselect("Seleccione la temporada:", TEMPORADAS,default="2025-1")

df_proyecto = df_total[df_total['Temporada'].isin(selector_temporada)]

# Inicializamos las listas para almacenar los DataFrames de cada temporada
df_UTI_list = []
df_REDI_list = []

# Iterar sobre cada temporada
for temporada in selector_temporada:
    # Construir la ruta al archivo UTI y REDI
    UTI = './' + temporada + '/UTI.xlsx'
    REDI = './' + temporada + '/REDI.xlsx'
    
    # Cargar los archivos correspondientes a la temporada
    df_UTI = pd.read_excel(UTI, sheet_name="Sheet1")
    df_REDI = pd.read_excel(REDI, sheet_name="Sheet1")
    
    # Reemplazar valores N/A (NaN) con ceros en el DataFrame de UTI
    df_UTI = df_UTI.fillna(0)
    
    # Almacenar los DataFrames en las listas
    df_UTI_list.append(df_UTI)
    df_REDI_list.append(df_REDI)

# Si necesitas combinar los DataFrames de todas las temporadas en uno solo, puedes usar pd.concat:
df_UTI = pd.concat(df_UTI_list, ignore_index=True)
df_REDI = pd.concat(df_REDI_list, ignore_index=True)

##################################################################################
#RATIOS GENERALES
df_acero_G= df_REDI.groupby(['Categoría'])['Peso(kg)'].sum().reset_index()
df_acero_G["Peso(kg)"] = df_acero_G["Peso(kg)"]/1000
df_acero_G.rename(columns={'Peso(kg)': 'Peso(Tn)'}, inplace=True)

df_soldadura_G= df_REDI[df_REDI['Desc.Corta'].str.startswith('SOLDADURA', na=False)]
df_soldadura_G= df_soldadura_G.groupby(['Categoría'])['Cantidad tomada'].sum().reset_index()
df_soldadura_G.rename(columns={'Cantidad tomada': 'Soldadura(kg)'}, inplace=True)

df_alambre_G= df_REDI[df_REDI['Desc.Corta'].str.startswith('ALAMBRE', na=False)]
df_alambre_G= df_alambre_G.groupby(['Categoría'])['Cantidad tomada'].sum().reset_index()
df_alambre_G.rename(columns={'Cantidad tomada': 'Alambre tub(kg)'}, inplace=True)

df_oxigeno_G= df_REDI[df_REDI['Desc.Corta'].isin(['OXIGENO IND.'])]
df_oxigeno_G= df_oxigeno_G.groupby(['Categoría'])['Cantidad tomada'].sum().reset_index()
df_oxigeno_G.rename(columns={'Cantidad tomada': 'Oxigeno(m3)'}, inplace=True)

df_disco_G= df_REDI[df_REDI['Desc.Corta'].str.startswith('DISCO', na=False)]
df_disco_G= df_disco_G.groupby(['Categoría'])['Cantidad tomada'].sum().reset_index()
df_disco_G.rename(columns={'Cantidad tomada': 'Discos(pz)'}, inplace=True)

df_ratio = pd.merge(df_acero_G, df_soldadura_G, on=['Categoría'], how='outer')
df_ratio = pd.merge(df_ratio, df_alambre_G, on=['Categoría'], how='outer')
df_ratio = pd.merge(df_ratio, df_oxigeno_G, on=['Categoría'], how='outer')
df_ratio = pd.merge(df_ratio, df_disco_G, on=['Categoría'], how='outer')
df_ratio.fillna(0, inplace=True)

df_ratio['Soldadura Total(kg)'] = (df_ratio['Soldadura(kg)']+df_ratio['Alambre tub(kg)']*1.67)

df_ratio['SoldxAcero'] = (df_ratio['Soldadura Total(kg)'])/df_ratio['Peso(Tn)']
df_ratio['OxigenoxAcero'] = df_ratio['Oxigeno(m3)']/df_ratio['Peso(Tn)']
df_ratio['DiscoxAcero'] = df_ratio['Discos(pz)']/df_ratio['Peso(Tn)']
df_ratio.fillna(0, inplace=True)

#Se elimina la categoría de : "PG" , "SISTEMAS AUXILIARES"
df_ratio = df_ratio[df_ratio['Categoría'].isin(['CASCO','ADITAMENTO','PANGA'])]

selector_categoria = st.sidebar.selectbox("Seleccione categoria:", ['CASCO','ADITAMENTO','PANGA'])

#Se filtra la categoría
df_ratio_filtrado = df_ratio[df_ratio['Categoría'].isin([selector_categoria])]
#df_ratio_filtrado = df_ratio_filtrado.query("`Peso(Tn)` > 0")

# Seleccionar columnas específicas
columnas_seleccionadas = ["Categoría", "SoldxAcero", "OxigenoxAcero","DiscoxAcero"]
df_seleccionado = df_ratio_filtrado[columnas_seleccionadas]
df_seleccionado.fillna(0, inplace=True)
print("df_ratio")
print(df_REDI)

# Configuración de la aplicación
st.title("📉 Ratios: "+ selector_categoria)

# Opciones configurables
def generate_gauge_options(column_name, value,tooltip):
    return {
        "tooltip": {
            "formatter": f"{tooltip}",
        },
        "series": [
            {
                "name": column_name,
                "type": "gauge",
                "progress": {"show": True, "width": 10},
                "axisLine": {
                    "lineStyle": {
                        "width": 10,
 
                    }
                },
                "pointer": {"width": 5},
                "title": {
                    "show": True,
                    "offsetCenter": [0, "70%"],
                    "fontSize": 20,
                    "text": column_name,
                },
                "detail": {
                    "valueAnimation": True,
                    "formatter": "{value}",
                    "fontSize": 30,
                },
                "data": [{"value": value, "name": column_name}],
            }
        ],
    }


col1, col2, col3 = st.columns(3)

with col1:
    with st.container(border=True):
        options = generate_gauge_options("SoldxAcero", int(df_seleccionado["SoldxAcero"].mean()),"Soldadura(kg) x Acero(Tn)")
        st_echarts(options, width="400px") 

with col2:
    with st.container(border=True):
        options = generate_gauge_options("OxigenoxAcero", int(df_seleccionado["OxigenoxAcero"].mean()),"Oxígeno(m3) x Acero(Tn)")
        st_echarts(options, width="400px")

with col3:
    with st.container(border=True):
        options = generate_gauge_options("DiscoxAcero", int(df_seleccionado["DiscoxAcero"].mean()),"Discos(pz) x Acero(Tn)")
        st_echarts(options, width="400px")
        
print(df_ratio)

col1,col2,col3,col4,col5= st.columns(5)
with col1:
    with st.container(border=True):
        st.metric(label="Peso(Tn)", value=f"{int(df_ratio_filtrado['Peso(Tn)'].sum())} Tn")
with col2:       
    with st.container(border=True):
        st.metric(label="Soldadura(kg)", value=f"{int(df_ratio_filtrado['Soldadura(kg)'].sum())} kg")
with col3:      
    with st.container(border=True):
        st.metric(label="Alambre tub(kg)", value=f"{int(df_ratio_filtrado['Alambre tub(kg)'].sum())} kg")
with col4:      
    with st.container(border=True):
        st.metric(label="Oxigeno(m3)", value=f"{int(df_ratio_filtrado['Oxigeno(m3)'].sum())} m3")
with col5:      
    with st.container(border=True):
        st.metric(label="Discos(pz)", value=f"{int(df_ratio_filtrado['Discos(pz)'].sum())} pz")
        
        


df_consolidado = df_REDI

REDI_filtrado =  df_REDI[df_REDI["Proyecto"].isin(df_proyecto["Proyecto"])]

# Agrupar por 'Material' y sumar 'Cantidad'
consolidados =  df_REDI.groupby(["Material",'Desc.Corta'], as_index=False).agg({'Cantidad': 'sum',"MAT Estimado": 'sum'})

#MAQUINARIA EN PROYECTOS
PD_MAQUINARIA = df_UTI[df_UTI['Liquidación'] > 0]

PD_GRUA= PD_MAQUINARIA[PD_MAQUINARIA['Denom.Operación'].str.contains('grua|grúa', case=False, na=False)]
PD_GRUA= PD_GRUA.groupby(['Proyecto'])['MOD'].sum().reset_index()
PD_GRUA.rename(columns={'MOD': 'GRUA'}, inplace=True)

PD_MONTACARGA= PD_MAQUINARIA[PD_MAQUINARIA['Denom.Operación'].str.contains('montacarga', case=False, na=False)]
PD_MONTACARGA= PD_MONTACARGA.groupby(['Proyecto'])['MOD'].sum().reset_index()
PD_MONTACARGA.rename(columns={'MOD': 'MONTACARGA'}, inplace=True)

df_maquinaria = pd.merge(PD_MONTACARGA, PD_GRUA, on='Proyecto', how='outer')
df_maquinaria.fillna(0, inplace=True)  

#LIMPIEZA CARRIL

PD_CARRIL = df_UTI[df_UTI['Liquidación'] > 0]

PD_CARRIL= PD_CARRIL[PD_CARRIL['Denom.Operación'].str.contains('limpieza de carril', case=False, na=False)]
PD_CARRIL= PD_CARRIL.groupby(['Proyecto'])['MOD'].sum().reset_index()
PD_CARRIL.rename(columns={'MOD': 'Limpieza de carril'}, inplace=True)
    
#ESFUERZO ADICIONAL
BD_ESFUERZO =  df_UTI[df_UTI['Liquidación'] > 0]
BD_ESFUERZO = BD_ESFUERZO[BD_ESFUERZO['Denom.Operación'].str.contains('esfuerzo', case=False, na=False)]
BD_ESFUERZO = BD_ESFUERZO.groupby(['Proyecto'])['MOD'].sum().reset_index()
BD_ESFUERZO.rename(columns={'MOD': 'Esfuerzo Adicional'}, inplace=True)


st.subheader("Consolidado de la temporada")
col1,col2,col3,col4= st.columns(4)
with col1:
    with st.container(border=True):
        st.metric(label="Grúa", value=f"S/ {df_maquinaria['GRUA'].sum():,.2f}")
with col2:       
    with st.container(border=True):
        st.metric(label="Montacarga", value=f"S/ {df_maquinaria['MONTACARGA'].sum():,.2f}")
with col3:      
    with st.container(border=True):
        st.metric(label="Limpieza de carril", value=f"S/ {PD_CARRIL['Limpieza de carril'].sum():,.2f}")
with col4:      
    with st.container(border=True):
        st.metric(label="Esfuerzo Adicional", value=f"S/ {BD_ESFUERZO['Esfuerzo Adicional'].sum():,.2f}")



#TABLA DE MATERIALES EMPLEADOS EN LA TEMPORADA
df_ratio = df_ratio[df_ratio['Categoría'].isin(['CASCO','ADITAMENTO','PANGA'])]

with st.expander("Materiales totales usados por temporada"):
    material = st.text_input("Ingrese busqueda")

    # Filtrar los datos según la entrada del usuario
    if material.strip():  # Verificar que el usuario ha ingresado texto
        filtrado = consolidados[consolidados['Desc.Corta'].str.contains(material, case=False, na=False)]
    else:
        filtrado = consolidados
        filtrado['Material'] = filtrado['Material'].astype(str)
        
    st.dataframe(filtrado,use_container_width=True, hide_index=True,column_config={
                            'MAT Estimado': st.column_config.NumberColumn(
                                "MAT Estimado",
                                format="S/ %.2f",  
                            )
                        })
